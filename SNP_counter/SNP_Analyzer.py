import operator
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord
from Bio.Alphabet import generic_dna
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, Series

"""
My modules
"""
from SNP_Functions import analysis

"""
End of my modules
"""
# TODO: Tidy it up
# TODO: Make an argument parser for Power Shell, CMD
input_file = "rasmus_rc.fasta"
try:
    input_file
except NameError:
    input_file = "test.fas"


def _SeqGap_Removal(seq):
    '''
    Takes a SeqRecord (FASTA) object and removes gaps.

    Arguments: SeqRecord (FASTA)
    Returns: SeqRecord (FASTA) - without gaps
    '''
    seq = seq.tomutable()
    while "-" in seq:
        seq.remove("-")
    seq = seq.toseq()
    return seq


def Biopython_Translatefix(seq):
    '''
    Takes a SeqRecord (FASTA - DNA). Removes gaps and makes the sequence dividable with 3 without remainder.
    Adds undefined Nucleotides at the end if necessary.

    Arguments: SeqRecord (FASTA)
    Returns: SeqRecord (FASTA) - without gaps + "N" / "NN" if needed.
    '''
    seq = seq.tomutable()
    while "-" in seq:
        seq.remove("-")
    seq = seq.toseq()
    if len(seq) % 3 == 0:
        return seq
    elif len(seq) % 3 == 1:
        return seq + "NN"
    else:
        return seq + "N"


def _SNP_Data_Recorder(Local_Report, alignment_position, SeqA_Position, SeqB_Position, SeqA_Nucleotide, SeqB_Nucleotide):
    '''
    Records SNP data to a dictionary passed to it (both target dictionary and data).

    Arguments: Target Dictionary; Alignment position; Position in Sequence A; Position is Sequence B;
                Nucleotide in Sequence A; Nucleotide in Sequence B
    '''
    Local_Report["Alignment_Position"].append(alignment_position)
    Local_Report["SeqA_Position"].append(SeqA_Position)
    Local_Report["SeqB_Position"].append(SeqB_Position)
    Local_Report["SeqA_Nucleotide"].append(SeqA_Nucleotide)
    Local_Report["SeqB_Nucleotide"].append(SeqB_Nucleotide)


def _Local_Report_Popper(Local_Report, index=0):
    '''
    Pops the first elements of the dictionary passed to it. (Alignment position; Position in Sequence A;
    Position is Sequence B; Nucleotide in Sequence A; Nucleotide in Sequence B)

    Arguments: Dictionary to process, index (optional)
    '''
    Local_Report["Alignment_Position"].pop(index)
    Local_Report["SeqA_Position"].pop(index)
    Local_Report["SeqB_Position"].pop(index)
    Local_Report["SeqA_Nucleotide"].pop(index)
    Local_Report["SeqB_Nucleotide"].pop(index)


def SNP_Registrator(seqA, seqB):
    '''
    Takes in two sequences. One reference sequence and a target sequence. Records SNPs between the two sequences.
    Also records gaps, "double gaps".

    Arguments: Reference sequence, target sequence.

    Returns: Dictionary, with all the SNP data.
    '''
    # TODO: Do I want to keep the print functions in comment?!
    gap_once = 0
    SNP_count = 0
    SeqA_Position = 0
    SeqB_Position = 0
    alignment_position = 1
    Local_Report = {"Alignment_Position": [], "SeqA_Position": [],
                    "SeqA_Nucleotide": [], "SeqB_Position": [], "SeqB_Nucleotide": []}
    for r in range(0, len(seqA)):
        # This if branch is optional for Multiple alignements, I left it
        # here,because it doesn'T do anything for paiwaise, beacuse of it's
        # nature / properties
        # TODO: IT has to be fixed here - I THINK IT is fixed - Have to come back later
        if seqA[r] == "-" and seqB[r] == "-":
            if gap_once != "both":
                SNP_count += 1
                # print(alignment_position, seqA[r], seqB[r], SNP_count)
                _SNP_Data_Recorder(Local_Report, alignment_position, str(SeqA_Position) +
                                   "+", str(SeqB_Position) + "+", seqA[r], seqB[r])
            gap_once = "both"
        elif seqA[r] != seqB[r]:
            if seqA[r] != "-" and seqB[r] != "-":
                SeqA_Position += 1
                SeqB_Position += 1
                SNP_count += 1
                gap_once = 0
                # print(alignment_position,  seqA[r], seqB[r], SNP_count)
                _SNP_Data_Recorder(Local_Report, alignment_position,
                                   str(SeqA_Position), str(SeqB_Position), seqA[r], seqB[r])
            elif seqA[r] != "-" and seqB[r] == "-" and gap_once != "B":
                gap_once = "B"
                SeqA_Position += 1
                SNP_count += 1
                # print(alignment_position,  seqA[r], seqB[r], SNP_count)
                _SNP_Data_Recorder(Local_Report, alignment_position,
                                   str(SeqA_Position), str(SeqB_Position) + "+", seqA[r], seqB[r])
            elif seqA[r] == "-" and seqB[r] != "-" and gap_once != "A":
                gap_once = "A"
                SeqB_Position += 1
                SNP_count += 1
                # print(alignment_position,  seqA[r], seqB[r], SNP_count)
                _SNP_Data_Recorder(Local_Report, alignment_position, str(
                    SeqA_Position) + "+", str(SeqB_Position), seqA[r], seqB[r])
        else:
            SeqA_Position += 1
            SeqB_Position += 1
        alignment_position += 1
        Local_Report["SNP_count"] = SNP_count
    # Asessing local mutation rate
    Local_Report["Mutation_Rate"] = (SNP_count / len(seqA)) * 100
    Local_Report["SeqA_Length"] = len(seqA)
    if Local_Report["SeqA_Nucleotide"][-1] == "-" and Local_Report["SeqB_Nucleotide"][-1] == "-":
        _Local_Report_Popper(Local_Report, -1)
    if Local_Report["SeqA_Nucleotide"][0] == "-" and Local_Report["SeqB_Nucleotide"][0] == "-":
        _Local_Report_Popper(Local_Report, 0)
    return Local_Report

# SNP Type Data recording happens in these functions


def _SNP_Type_Increment(SNP_Type, SNP_Type_Data):
    '''
    Adds 1 to the sum of the given type of SNP in a dictionary designed to collect the number of each SNP type.

    Arguments: Type of SNP in the format of "NN". Where the first letter is Nucleotide in sequence A (or gap)
    and the second letter is Nucelotide in Sequence B (or gap).; Dictionary that counts the number of each SNP.
    '''
    SNP_Type_Data[SNP_Type] += 1


def _SNP_Type_Data_Establisher(string, Local_SNP_Type_Data):
    '''
    Builds up a Dictionary to record SNP occurence numbers, starting with 0 for each SNP type.

    Arguments: String (I made it this way, so it's reusable for proteins later)
                that contains all possible characters for AA or bases (DNA); Dictionary to build.
    '''
    Local_SNP_Type_Data["--"] = 0
    for l in string:
        generated_key_base = l
        for l in string:
            generated_key = generated_key_base
            if generated_key == l:
                pass
            else:
                generated_key += l
                Local_SNP_Type_Data[generated_key] = 0
    return Local_SNP_Type_Data


def SNP_Type_Data_Generator(SNP_Report):
    '''
    Takes the SNP Report Dictionary, with the collection of SNPs, counts each SNP type and returns a dictionary
    with the types and values.

    Arguments:
        Dictionary with the SNP data - (SNP_Report).

    Returns:
        Dictionary with the number of each SNP type occured for each target sequence.
    '''
    SNP_Type_Data = {}
    for key in SNP_Report:
        Local_SNP_Type_Data = {}
        Local_SNP_Type_Data = _SNP_Type_Data_Establisher("AGTC-", Local_SNP_Type_Data)
        for r in range(0, len(SNP_Report[key]["Alignment_Position"])):
            SNP_Type = SNP_Report[key]["SeqA_Nucleotide"][r] + SNP_Report[key]["SeqB_Nucleotide"][r]
            _SNP_Type_Increment(SNP_Type, Local_SNP_Type_Data)
            # print(SNP_Type) - I have it for Debugging purposes
        SNP_Type_Data[key] = Local_SNP_Type_Data
        SNP_Type_Data[key]["Mutation_Rate"] = SNP_Report[key]["Mutation_Rate"]
        SNP_Type_Data[key]["SeqA_Length"] = SNP_Report[key]["SeqA_Length"]
    return SNP_Type_Data


def SNP_Data_Summer(SNP_Type_Data):
    '''
    Sums up the number of SNPs trhough out the whole run.

    Arguments:
        Dictionary with SNP types and their numbers (SNP_Type_Data)

    Returns:
        Dictionary with the summed up numbers of SNPs for each type.
    '''
    SNP_Type_Data_Sum = {}
    SNP_Type_Data_Sum = _SNP_Type_Data_Establisher("ATGC-", SNP_Type_Data_Sum)
    SNP_Type_Data_Sum["Mutation_Rate"] = 0
    Summed_Sequence_Length = 0
    for key_lvl1 in SNP_Type_Data:
        for key_lvl2 in SNP_Type_Data[key_lvl1]:
            if key_lvl2 != "Mutation_Rate" and key_lvl2 != "SeqA_Length":
                SNP_Type_Data_Sum[key_lvl2] += SNP_Type_Data[key_lvl1][key_lvl2]
        SNP_Type_Data_Sum["Mutation_Rate"] += SNP_Type_Data[key_lvl1]["Mutation_Rate"] * \
            SNP_Type_Data[key_lvl1]["SeqA_Length"]
        Summed_Sequence_Length += SNP_Type_Data[key_lvl1]["SeqA_Length"]
    SNP_Type_Data_Sum["Mutation_Rate"] = SNP_Type_Data_Sum["Mutation_Rate"] / Summed_Sequence_Length
    return SNP_Type_Data_Sum


# Sofware starts here


# Get the iteratable object
record_iterator = SeqIO.parse(input_file, "fasta")

# Setting variables for the upcoming while loop
w = 1
SNP_Report = {}
sequences = []
# TODO: Multiple References have to be implemented
reference_record = next(record_iterator)
reference_record.seq.alphabet = generic_dna
print (reference_record)
# Append Reference to the 1st item!!!
sequences.insert(0, SeqRecord(_SeqGap_Removal(
    reference_record.seq), reference_record.id))

# Registration of SNPs and InDels into a Dictionary
while True:
    # Getting the SeqRecords to variables, so sequences and their ID will be
    # accessable
    subject_record = next(record_iterator)
    subject_record.seq.alphabet = generic_dna
    for key in SNP_Report:
        if key == subject_record.id:
            subject_record.id += "-RiD"
    # Running the registrator function on the two given sequences. (Ref+Subject)
    SNP_Report[str(subject_record.id)] = SNP_Registrator(
        reference_record.seq, subject_record.seq)
    sequences.append(SeqRecord(_SeqGap_Removal(
        subject_record.seq), subject_record.id))
    print("You have a total of " + str(SNP_Report[str(subject_record.id)]["SNP_count"]) +
          " number SNPs, including InDels. In this sequence: " + str(subject_record.id))
    reference_record = next(record_iterator, None)
    if reference_record is not None:
        reference_record.seq.alphabet = generic_dna
    elif reference_record is None:
        break

# FASTA OUTPUT
# It writes the Seq Records (only ID + Seq) to Fasta format
SeqIO.write(sequences, "OutputTest.fas", "fasta")

# Check the dictionary for better DATA > number of nucleotide transitions.
SNP_Type_Data = SNP_Type_Data_Generator(SNP_Report)
SNP_Type_Data_Sum = SNP_Data_Summer(SNP_Type_Data)
'''
Debugging purposes - Doesn't show up in Excle or FASTA files
'''
print(SNP_Type_Data)
print(SNP_Type_Data_Sum)
print(SNP_Report)


# EXCEL OUTPUT - Setting up the sheets
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SNP Report"
ws.sheet_properties.tabColor = "00FF00"

ws_TypeData = wb.create_sheet("SNP Types", 1)
ws_TypeData.sheet_properties.tabColor = "0000FF"

ws_ChartData = wb.create_sheet("Chart Data")
ws_ChartData.sheet_properties.tabColor = "FFFF00"

# Styling for the Excel sheet
borderstyle = Side(style="thin", color="000000")

highlight = NamedStyle(name="highlight")
highlight.border = Border(top=borderstyle, bottom=borderstyle, right=borderstyle)
highlight.fill = PatternFill("solid", fgColor="EEEEEE")
highlight.font = Font(bold=True)

wb.add_named_style(highlight)
# Styling the Mutation rate cells
Red_Highlight = NamedStyle(name="Red_Highlight")
Red_Highlight.number_format = "0.00E+00"
Red_Highlight.font = Font(bold=True, italic=True, size=12, color="FF0000")
Red_Highlight.alignment = Alignment(horizontal="right")
Red_Highlight.fill = PatternFill("solid", fgColor="EEEEEE")
Red_Highlight.border = Border(top=borderstyle, bottom=borderstyle, right=borderstyle)

Left_Highlight = NamedStyle(name="Left_Highlight")
Left_Highlight.font = Font(bold=True, italic=True, size=12, color="0000FF")
Left_Highlight.alignment = Alignment(horizontal="left")

Left_Mutation = NamedStyle(name="Left_Mutation")
Left_Mutation.font = Font(bold=True, italic=True, size=12, color="0000FF")
Left_Mutation.alignment.alignment = Alignment(horizontal="left")
Left_Mutation.number_format = "0.00E+00"

# Saving the SNP_Report Dictionary to a worksheet (positions, and nucelotide differences)


def SNP_Report_To_Excel(worksheet, SNP_Report):
    '''
    Takes a Worskheet in the virtual excel file and the SNP report dictionary. Makes the worksheet nice and neat (check excel file)
    Containing all the information about each SNP.

    Arguments:
        Worksheet of the virtual excel file.

        Dictionary with the SNP data to transfer this data to excel.
    '''
    r = 1
    for keys, value in sorted(SNP_Report.items()):
        c = 1
        worksheet.cell(row=r, column=1, value=keys).font = Font(bold=True, italic=True, size=14)
        worksheet.merge_cells("B" + str(r) + ":C" + str(r))
        worksheet.cell(row=r, column=2, value="#SNPs+InDel:").font = Font(bold=True, italic=True, size=12)
        worksheet.cell(row=r, column=4, value=SNP_Report[keys]["SNP_count"]).style = Left_Highlight
        worksheet.merge_cells("E" + str(r) + ":G" + str(r))
        worksheet.merge_cells("H" + str(r) + ":I" + str(r))
        worksheet.cell(row=r, column=5, value="Mutation rate [%]:").font = Font(bold=True, italic=True, size=12)
        worksheet.cell(row=r, column=8, value=SNP_Report[keys]["Mutation_Rate"]).style = Left_Mutation
        worksheet.cell(row=r + 1, column=c, value="Alignment Postion").style = highlight
        worksheet.cell(row=r + 2, column=c, value="Reference Sequence Position")
        worksheet.cell(row=r + 3, column=c, value="Reference Nucleotide").style = highlight
        worksheet.cell(row=r + 4, column=c, value=str(keys) + " Sequence Poisiton")
        worksheet.cell(row=r + 5, column=c, value=str(keys) + " Nucleotide").style = highlight
        i = 0
        c = 2
        for item in SNP_Report[keys]["Alignment_Position"]:
            worksheet.cell(row=r + 1, column=c, value=SNP_Report[keys]["Alignment_Position"][i]).style = highlight
            worksheet.cell(row=r + 2, column=c, value=SNP_Report[keys]["SeqA_Position"][i])
            worksheet.cell(row=r + 3, column=c, value=SNP_Report[keys]["SeqA_Nucleotide"][i]).style = highlight
            worksheet.cell(row=r + 4, column=c, value=SNP_Report[keys]["SeqB_Position"][i])
            worksheet.cell(row=r + 5, column=c, value=SNP_Report[keys]["SeqB_Nucleotide"][i]).style = highlight
            c += 1
            i += 1
        r += 6
    worksheet.column_dimensions["A"].width = 30


# Saving the Types
# TODO: Make merged cells, and represent the percentage of each type
def SNP_Type_To_Excel(worksheet, worksheet_chart_data, SNP_Type_Data, SNP_Type_Data_Sum):
    '''
    Takes a Worksheet to record date about SNP types and another Worksheet to make a pie-chart of the Summed Data.

    Arguments:
        Worksheet to record SNP types.

        Worksheet to construct chart in.

        Dictionary with all the SNP types and the data.

        Dictionary of the summed SNP types.
    '''
    worksheet.column_dimensions["A"].width = 20
    c = 1
    r = 1
    # First column > Names of Types of SNPs
    for key, value in sorted(SNP_Type_Data_Sum.items()):
        if key != "Mutation_Rate" and key != "SeqA_Length":
            worksheet.cell(row=r + 1, column=1, value=key)
            if r % 2 == 1:
                worksheet.cell(row=r + 1, column=1).style = highlight
            r += 1
    worksheet.cell(row=r + 1, column=1, value="Sum").font = Font(bold=True)
    worksheet.cell(row=r + 2, column=1, value="Mutation Rate[%]").style = highlight
    # Registration of each subject sequence
    for key_lvl1, value_lvl1 in sorted(SNP_Type_Data.items()):
        r = 1
        worksheet.cell(row=1, column=c + 1, value=key_lvl1).font = Font(bold=True, italic=True)
        # TODO: IT needs fixing
        # w = len(worksheet.cell(row=1, column=c + 1).value) + 1
        w = 15
        if w >= 12:
            worksheet.column_dimensions[get_column_letter(c + 1)].width = w
        else:
            worksheet.column_dimensions[get_column_letter(c + 1)].width = 12
        SNP_Type_Sum = 0
        for key_lvl2, value_lvl2 in sorted(SNP_Type_Data[key_lvl1].items()):
            if key_lvl2 != "Mutation_Rate" and key_lvl2 != "SeqA_Length":
                worksheet.cell(row=r + 1, column=c + 1, value=value_lvl2)
                SNP_Type_Sum += value_lvl2
                if r % 2 == 1:
                    worksheet.cell(row=r + 1, column=c + 1).style = highlight
                else:
                    worksheet.cell(row=r + 1, column=c + 1).font = Font(bold=True)
                r += 1
        worksheet.cell(row=r + 1, column=c + 1, value=SNP_Type_Sum).font = Font(bold=True)
        worksheet.cell(row=r + 2, column=c + 1, value=SNP_Type_Data[key_lvl1]["Mutation_Rate"]).style = Red_Highlight
        c += 1
    r = 1
    # Registration of The Sum SNPs
    worksheet.cell(row=1, column=c + 1, value="SNP Sum").font = Font(bold=True, italic=True)
    w = len(worksheet.cell(row=1, column=c + 1).value) + 1
    if w >= 12:
        worksheet.column_dimensions[get_column_letter(c + 1)].width = w
    else:
        worksheet.column_dimensions[get_column_letter(c + 1)].width = 12
    for key, value in sorted(SNP_Type_Data_Sum.items()):
        if key != "Mutation_Rate":
            worksheet.cell(row=r + 1, column=c + 1, value=value)
            if r % 2 == 1:
                worksheet.cell(row=r + 1, column=c + 1).style = highlight
            else:
                worksheet.cell(row=r + 1, column=c + 1).font = Font(bold=True)
            SNP_Type_Sum += value
            r += 1
        worksheet.cell(row=r + 1, column=c + 1, value=SNP_Type_Sum).font = Font(bold=True)
    worksheet.cell(row=r + 2, column=c + 1, value=SNP_Type_Data_Sum["Mutation_Rate"]).style = Red_Highlight
    # Chart Making starts here
    # Sorting the Data in descending order, to obtian nice chart
    r = 1
    for key, value in sorted(SNP_Type_Data_Sum.items(), key=operator.itemgetter(1), reverse=True):
        if key != "Mutation_Rate" and key != "SeqA_Length":
            worksheet_chart_data.cell(row=r, column=1, value=key)
            worksheet_chart_data.cell(row=r, column=2, value=value)
            r += 1

    # Creating the chart itself, from the newly sorted data
    cs_TypeData = wb.create_sheet("SNP Distribution", 2)
    
    chart = PieChart()
    labels = Reference(worksheet_chart_data, min_col=1, min_row=2, max_row=21)
    data = Reference(worksheet_chart_data, min_col=2, min_row=1, max_row=21)
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(labels)
    #chart.series = (Series(data),)
    chart.title = "PieChart of SNP-Type distribution"

    cs_TypeData.add_chart(chart, "A1")
    
    
"""   
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Pies sold by category"
"""


# BIG MAGIC - Setting each worksheet
SNP_Report_To_Excel(ws, SNP_Report)
SNP_Type_To_Excel(ws_TypeData, ws_ChartData, SNP_Type_Data, SNP_Type_Data_Sum)
# Saving the WorkBook from RAM to HDD
# TODO: Get an ouptput File name from ARG parser
wb.save(input_file.rsplit(".", 1)[0] + "ExcelResults.xlsx")


# This is the Purpose of TranslateFix function.
# TODO: This part will be removed before the hand-in.
'''
# Fixing the length of the sequences to be dividable with 3 and removing gaps
reference_record_fix = Biopython_Translatefix(reference_record.seq)
subject_record_fix = Biopython_Translatefix(subject_record.seq)


reference_record_AA = reference_record_fix.translate(stop_symbol="@")
subject_record_AA = subject_record_fix.translate(stop_symbol="@")

# Finding the first start codon
print(reference_record_AA.find("M"))


# You can use functions in indexing as long as it returns a number >>>
# #powerful
'''
