# -*- coding: utf-8 -*-
"""
Excel Handling functions for SNP counter
Verti 2018.02.28.
"""
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, Series

# Styling for the Excel sheet
borderstyle = Side(style="thin", color="000000")

highlight = NamedStyle(name="highlight")
highlight.border = Border(top=borderstyle, bottom=borderstyle, right=borderstyle)
highlight.fill = PatternFill("solid", fgColor="EEEEEE")
highlight.font = Font(bold=True)

#wb.add_named_style(highlight)

# Styling the Mutation rate cells
Red_Highlight = NamedStyle(name="Red_Highlight")
Red_Highlight.number_format = "0.00E+00"
Red_Highlight.font = Font(bold=True, italic=True, size=12, color="FF0000")
Red_Highlight.alignment = Alignment(horizontal="right")
Red_Highlight.fill = PatternFill("solid", fgColor="EEEEEE")
Red_Highlight.border = Border(top=borderstyle, bottom=borderstyle, right=borderstyle)

Left_Highlight = NamedStyle(name="Left_Highlight")
Left_Highlight.font = Font(bold=True, italic=True, size=12, color="0000FF")
Left_Highlight.alignment = Alignment(horizontal="left")

Left_Mutation = NamedStyle(name="Left_Mutation")
Left_Mutation.font = Font(bold=True, italic=True, size=12, color="0000FF")
Left_Mutation.alignment.alignment = Alignment(horizontal="left")
Left_Mutation.number_format = "0.00E+00"


# Saving the SNP_Report Dictionary to a worksheet (positions, and nucelotide differences)
def SNP_Report_To_Excel(worksheet, SNP_Report):
    '''
    Takes a Worskheet in the virtual excel file and the SNP report dictionary. Makes the worksheet nice and neat (check excel file)
    Containing all the information about each SNP.

    Arguments:
        Worksheet of the virtual excel file.

        Dictionary with the SNP data to transfer this data to excel.
    '''
    r = 1
    for keys, value in sorted(SNP_Report.items()):
        c = 1
        worksheet.cell(row=r, column=1, value=keys).font = Font(bold=True, italic=True, size=14)
        worksheet.merge_cells("B" + str(r) + ":C" + str(r))
        worksheet.cell(row=r, column=2, value="#SNPs+InDel:").font = Font(bold=True, italic=True, size=12)
        worksheet.cell(row=r, column=4, value=SNP_Report[keys]["SNP_count"]).style = Left_Highlight
        worksheet.merge_cells("E" + str(r) + ":G" + str(r))
        worksheet.merge_cells("H" + str(r) + ":I" + str(r))
        worksheet.cell(row=r, column=5, value="Mutation rate [%]:").font = Font(bold=True, italic=True, size=12)
        worksheet.cell(row=r, column=8, value=SNP_Report[keys]["Mutation_Rate"]).style = Left_Mutation
        worksheet.cell(row=r + 1, column=c, value="Alignment Postion").style = highlight
        worksheet.cell(row=r + 2, column=c, value="Reference Sequence Position")
        worksheet.cell(row=r + 3, column=c, value="Reference Nucleotide").style = highlight
        worksheet.cell(row=r + 4, column=c, value=str(keys) + " Sequence Poisiton")
        worksheet.cell(row=r + 5, column=c, value=str(keys) + " Nucleotide").style = highlight
        i = 0
        c = 2
        for item in SNP_Report[keys]["Alignment_Position"]:
            worksheet.cell(row=r + 1, column=c, value=SNP_Report[keys]["Alignment_Position"][i]).style = highlight
            worksheet.cell(row=r + 2, column=c, value=SNP_Report[keys]["SeqA_Position"][i])
            worksheet.cell(row=r + 3, column=c, value=SNP_Report[keys]["SeqA_Nucleotide"][i]).style = highlight
            worksheet.cell(row=r + 4, column=c, value=SNP_Report[keys]["SeqB_Position"][i])
            worksheet.cell(row=r + 5, column=c, value=SNP_Report[keys]["SeqB_Nucleotide"][i]).style = highlight
            c += 1
            i += 1
        r += 6
    worksheet.column_dimensions["A"].width = 30
    return worksheet